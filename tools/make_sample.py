# -*- coding: utf-8 -*-
"""動作確認用のダミーシフトExcelを生成する。

実物の「7月シフト.xlsx」と同じシート構造(「2）ｼﾌﾄﾊﾟﾀｰﾝ」「3）入力用」)を
ダミー氏名で再現する。実物ファイルは一切読まない。

usage: python make_sample.py [出力パス]  (default: sample/sample.xlsx)
requires: openpyxl
"""
import datetime as dt
import random
import sys
from pathlib import Path

import openpyxl

YEAR, MONTH = 2025, 7
DAYS = 31

# パターン名: (始業, 終業, 休憩, 表示, 分類)  分類: 1=開番 2=閉番 3=中番 0=その他
PATTERNS = {
    "早":     ("09:30", "19:00", "01:30", "早",    1),
    "超早":   ("08:30", "18:00", "01:30", "超早",  1),
    "微早":   ("09:15", "18:45", "01:30", "微早",  1),
    "早②":   ("09:30", "19:00", "01:00", "早②",  1),
    "微早②": ("09:15", "18:45", "01:00", "微早②", 1),
    "早17":   ("09:30", "17:00", "01:00", "早17",  1),
    "遅":     ("11:15", "20:45", "01:30", "遅",    2),
    "遅②":   ("11:15", "20:45", "01:00", "遅②",  2),
    "遅金":   ("11:45", "21:15", "01:30", "遅金",  2),
    "遅金②": ("11:45", "21:15", "01:00", "遅金②", 2),
    "14L":    ("14:00", "21:00", "01:00", "14L",   2),
    "16L":    ("16:00", "21:00", "00:30", "16L",   2),
    "1215L":  ("12:15", "20:45", "01:00", "1215L", 2),
    "1419":   ("14:00", "19:00", "00:00", "1419",  3),
    "1217":   ("12:00", "17:00", "00:00", "1217",  3),
    "1116":   ("11:00", "16:00", "00:00", "1116",  3),
    "会":     ("10:00", "17:30", "01:00", "会議",  0),
    "研":     ("10:00", "17:30", "01:00", "研修",  0),
    "有給":   ("09:30", "19:00", "01:30", "有給",  0),
}

# (CD, 役職, 氏名, 雇用区分, 部門)  氏名は全てダミー
STAFF = [
    (10001, "店長",       "山田 太郎", 1, "店長"),
    (10101, "部門ﾘｰﾀﾞｰ", "佐藤 花子", 1, "季節AV"),
    (10102, "",           "伊藤 三奈", 1, "季節AV"),
    (10103, "",           "渡辺 圭",   2, "季節AV"),
    (10201, "部門ﾘｰﾀﾞｰ", "鈴木 一郎", 1, "家電S"),
    (10202, "",           "中村 洋子", 2, "家電S"),
    (10203, "",           "小林 誠",   2, "家電S"),
    (10301, "部門ﾘｰﾀﾞｰ", "高橋 未来", 1, "情報S"),
    (10302, "",           "加藤 光",   2, "情報S"),
    (10401, "通信ﾘｰﾀﾞｰ", "田中 三郎", 1, "通信S"),
    (10402, "",           "吉田 恵",   2, "通信S"),
    (10403, "",           "山本 空",   2, "通信S"),
]

OPEN_CODES = ["早", "超早", "微早", "早②", "微早②"]
CLOSE_CODES = ["遅", "遅②", "遅金", "14L", "1215L"]
MID_CODES = ["1419", "1217", "1116"]


def build_shifts():
    """CD -> [day1..day31] のシフト記号(Noneは休み)。決定的に生成する。"""
    rng = random.Random(20250701)
    shifts = {}
    for cd, role, _name, emp, _dept in STAFF:
        row = []
        for d in range(1, DAYS + 1):
            wd = dt.date(YEAR, MONTH, d).weekday()
            # 週2日ペースの休み(人によってずらす)
            if (d + cd) % 7 in (0, 3):
                row.append(None)
                continue
            if role == "店長":
                row.append("超早" if wd in (0, 4) else "早")
            elif role:  # リーダー: 人ごとに位相をずらして開け・閉めに散らばす
                pool = OPEN_CODES if (cd // 100 + d) % 2 == 0 else CLOSE_CODES
                row.append(rng.choice(pool))
            elif emp == 1:
                row.append(rng.choice(OPEN_CODES + CLOSE_CODES))
            else:  # パート
                row.append(rng.choice(OPEN_CODES + CLOSE_CODES + MID_CODES))
        shifts[cd] = row
    # 警告検出テスト用の意図的な穴:
    # 7/15: 鍵保持者全員を閉め番以外にする → 閉めに鍵保持者ゼロ
    # 7/22: 鍵保持者全員を開け番以外にする → 開けに鍵保持者ゼロ
    leaders = [cd for cd, role, *_ in STAFF if role]
    for cd, role, *_ in STAFF:
        if role == "店長":
            shifts[cd][15 - 1] = "早"
            shifts[cd][22 - 1] = None
        elif role:
            shifts[cd][15 - 1] = "早②" if cd % 2 else None
            shifts[cd][22 - 1] = "遅②" if cd % 2 else None
    assert leaders
    return shifts


def t(s):
    h, m = s.split(":")
    return dt.time(int(h), int(m))


def main(out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- 2）ｼﾌﾄﾊﾟﾀｰﾝ ---
    ws = wb.create_sheet("2）ｼﾌﾄﾊﾟﾀｰﾝ")
    ws["B2"] = "シフトパターンシート"
    for c, h in zip("BCDEFGHIJ", ["ﾊﾟﾀｰﾝ", "始業", "終業", "休憩", "勤務\n時間", "実働", "表示", "", "1=開番\n2=閉番\n3=中番"]):
        ws[f"{c}4"] = h
    for i, (name, (start, end, brk, disp, kind)) in enumerate(PATTERNS.items()):
        r = 6 + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, name)
        ws.cell(r, 3, t(start))
        ws.cell(r, 4, t(end))
        ws.cell(r, 5, t(brk))
        ws.cell(r, 8, disp)
        ws.cell(r, 9, "→")
        ws.cell(r, 10, kind)

    # --- 3）入力用 ---
    ws = wb.create_sheet("3）入力用")
    ws["D1"] = "999"
    ws["E1"] = "サンプル店"
    # 11行目: 日付(datetime) J列以降
    for d in range(1, DAYS + 1):
        ws.cell(11, 9 + d, dt.datetime(YEAR, MONTH, d))
    # 12行目: ヘッダー
    for col, h in [(1, "行番号"), (4, "CD"), (5, "役職"), (6, "担当者名")]:
        ws.cell(12, col, h)
    shifts = build_shifts()
    r = 14
    prev_dept = None
    for cd, role, name, emp, dept in STAFF:
        if dept != prev_dept:
            r += 2  # 実物同様、部門間に隙間を空ける
            prev_dept = dept
        ws.cell(r, 4, cd)
        if role:
            ws.cell(r, 5, role)
        ws.cell(r, 6, name)
        ws.cell(r, 7, emp)
        for d in range(1, DAYS + 1):
            code = shifts[cd][d - 1]
            if code:
                ws.cell(r, 9 + d, code)
        r += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).resolve().parent.parent / "sample" / "sample.xlsx"
    main(out)
